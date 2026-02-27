import pandas as pd
import io
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(period_data, financial_metrics):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # --- Feuille 1 : Résumé ---
        summary_data = {
            'Période': [period_data.period_name],
            'Début': [period_data.start_date.strftime('%Y-%m-%d')],
            'Fin': [period_data.end_date.strftime('%Y-%m-%d')],
            'Jours': [len(period_data.daily_data)],
            'Opérations totales': [financial_metrics.period_summary['total_operations']],
            'Gains période ($)': [financial_metrics.period_gains],
            'Commission mensuelle ($)': [financial_metrics.your_commission_monthly]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Résumé', index=False, startrow=1)

        # Mise en forme de la feuille Résumé
        workbook = writer.book
        sheet = writer.sheets['Résumé']

        # Titre
        sheet['A1'] = 'RAPPORT OPSGAIN'
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells('A1:G1')
        sheet['A1'].alignment = Alignment(horizontal='center')

        # En-têtes de colonnes
        for col_num, value in enumerate(df_summary.columns, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = value
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        # Ajustement automatique des colonnes
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[col_letter].width = adjusted_width

        # --- Feuille 2 : Activité journalière ---
        period_data.daily_data.to_excel(writer, sheet_name='Activité journalière', index=False)
        sheet2 = writer.sheets['Activité journalière']
        # Mettre les en-têtes en gras
        for cell in sheet2[1]:
            cell.font = Font(bold=True)

        # --- Feuille 3 : Performance engins ---
        period_data.engins_data.to_excel(writer, sheet_name='Engins', index=False)
        sheet3 = writer.sheets['Engins']
        for cell in sheet3[1]:
            cell.font = Font(bold=True)

        # --- Feuille 4 : Opérations récentes ---
        period_data.recent_ops.to_excel(writer, sheet_name='Opérations récentes', index=False)
        sheet4 = writer.sheets['Opérations récentes']
        for cell in sheet4[1]:
            cell.font = Font(bold=True)

        # --- Feuille 5 : Détail des gains ---
        gains_df = pd.DataFrame([
            {'Type': 'Temps', 'Journalier ($)': financial_metrics.breakdown['time_gain'],
             'Période ($)': financial_metrics.breakdown['time_gain_period']},
            {'Type': 'Erreurs', 'Journalier ($)': financial_metrics.breakdown['error_gain'],
             'Période ($)': financial_metrics.breakdown['error_gain_period']},
            {'Type': 'Maintenance', 'Journalier ($)': financial_metrics.breakdown['maintenance_gain'],
             'Période ($)': financial_metrics.breakdown['maintenance_gain_period']},
            {'Type': 'Carburant', 'Journalier ($)': financial_metrics.breakdown['fuel_gain'],
             'Période ($)': financial_metrics.breakdown['fuel_gain_period']},
        ])
        gains_df.to_excel(writer, sheet_name='Détail gains', index=False)
        sheet5 = writer.sheets['Détail gains']
        for cell in sheet5[1]:
            cell.font = Font(bold=True)

        # Format monétaire pour les colonnes de gains
        for sheet in [sheet2, sheet3, sheet4, sheet5]:
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00 [$-409]'  # format monétaire dollar (adaptez)

    return output.getvalue()